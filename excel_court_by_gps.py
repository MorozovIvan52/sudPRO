#!/usr/bin/env python3
"""
Поиск суда по GPS и запись подробных сведений (45 полей) в Excel пользователя.

Вход: Excel-файл с колонками:
  - «Адрес» или «адрес» (текстовый адрес — будет геокодирован), либо
  - «Широта»/«Долгота» или «lat»/«lng» (координаты WGS84)
  - опционально: «Сумма долга» или «debt_amount» (для расчёта госпошлины)

Выход: тот же файл с добавленными 45 колонками (подробный список сведений по суду).
Либо новый файл: excel_court_by_gps.py input.xlsx -o result.xlsx

Запуск из корня проекта:
  python excel_court_by_gps.py путь/к/файлу.xlsx
  python excel_court_by_gps.py путь/к/файлу.xlsx -o результат.xlsx
"""
import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass


def _find_lat_lng_columns(header_row: list) -> tuple:
    """Возвращает (idx_lat, idx_lng) по заголовкам. Ищет широта/долгота, lat/lng."""
    header_lower = [str(h).strip().lower() if h is not None else "" for h in header_row]
    idx_lat = idx_lng = None
    for i, h in enumerate(header_lower):
        if h in ("lat", "широта", "latitude"):
            idx_lat = i
        if h in ("lng", "lon", "долгота", "longitude"):
            idx_lng = i
    return idx_lat, idx_lng


def _find_address_column(header_row: list) -> int:
    for i, h in enumerate(header_row):
        if h is None:
            continue
        s = str(h).strip().lower()
        if s in ("адрес", "address", "адрес регистрации"):
            return i
    return -1


def _find_debt_column(header_row: list) -> int:
    for i, h in enumerate(header_row):
        if h is None:
            continue
        s = str(h).strip().lower()
        if s in ("сумма долга", "debt_amount", "сумма", "долг"):
            return i
    return -1


def process_excel(input_path: str, output_path: str = None) -> str:
    try:
        import openpyxl
    except ImportError:
        print("Установите: pip install openpyxl")
        sys.exit(1)

    from court_locator.main import CourtLocator
    from court_locator.court_details import COURT_DETAIL_COLUMNS, build_court_details

    input_path = Path(input_path)
    if not input_path.exists():
        print("Файл не найден:", input_path)
        sys.exit(1)
    if output_path is None:
        output_path = input_path.parent / (input_path.stem + "_result" + input_path.suffix)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Лист пуст.")
        wb.close()
        return str(output_path)

    header_row = list(rows[0])
    idx_lat, idx_lng = _find_lat_lng_columns(header_row)
    idx_addr = _find_address_column(header_row)
    idx_debt = _find_debt_column(header_row)

    if idx_lat is None or idx_lng is None:
        if idx_addr < 0:
            print("В Excel нужны колонки «Широта» и «Долгота» (или lat/lng) либо «Адрес».")
            sys.exit(1)
        need_geocode = True
    else:
        need_geocode = False

    # Добавляем заголовки 45 колонок справа
    start_col = len(header_row) + 1
    for c, name in enumerate(COURT_DETAIL_COLUMNS, start=start_col):
        ws.cell(row=1, column=c, value=name)
    num_header_cols = start_col + len(COURT_DETAIL_COLUMNS) - 1

    locator = CourtLocator(use_cache=True)
    try:
        for row_idx, row in enumerate(rows[1:], start=2):
            lat, lng = None, None
            address_str = None
            debt_val = None

            if not need_geocode:
                if idx_lat is not None and idx_lat < len(row) and row[idx_lat] is not None:
                    try:
                        lat = float(row[idx_lat])
                    except (TypeError, ValueError):
                        pass
                if idx_lng is not None and idx_lng < len(row) and row[idx_lng] is not None:
                    try:
                        lng = float(row[idx_lng])
                    except (TypeError, ValueError):
                        pass
            if idx_addr >= 0 and idx_addr < len(row) and row[idx_addr]:
                address_str = str(row[idx_addr]).strip()

            if idx_debt >= 0 and idx_debt < len(row) and row[idx_debt] is not None:
                try:
                    debt_val = float(row[idx_debt])
                except (TypeError, ValueError):
                    pass

            court = None
            if lat is not None and lng is not None:
                court = locator.locate_court(lat=lat, lng=lng)
            if court is None and address_str:
                court = locator.locate_court(address=address_str)

            if court:
                details = build_court_details(
                    court,
                    normalized_address=address_str,
                    debt_amount=debt_val,
                )
                for c, col_name in enumerate(COURT_DETAIL_COLUMNS, start=start_col):
                    val = details.get(col_name, "")
                    ws.cell(row=row_idx, column=c, value=val)
            else:
                for c in range(start_col, num_header_cols + 1):
                    ws.cell(row=row_idx, column=c, value="")
    finally:
        locator.close()

    wb.save(output_path)
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description="Поиск суда по GPS по строкам Excel, запись 45 полей сведений.")
    parser.add_argument("excel_path", help="Путь к Excel-файлу от пользователя")
    parser.add_argument("-o", "--output", default=None, help="Путь к результирующему файлу (по умолчанию: имя_файла_result.xlsx)")
    args = parser.parse_args()
    out = process_excel(args.excel_path, args.output)
    print("Готово. Результат записан:", out)


if __name__ == "__main__":
    main()
